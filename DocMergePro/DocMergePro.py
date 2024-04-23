import os
from PyPDF2 import PdfFileMerger
from docx import Document
import openpyxl as px

# Function to merge PDF files into a single PDF
def merge_pdf_files(pdf_files, output_file):
    merger = PdfFileMerger()
    for pdf_file in pdf_files:
        merger.append(pdf_file)
    merger.write(output_file)
    merger.close()

# Function to merge Word files into a single Word document
def merge_word_files(word_files, output_file):
    merged_doc = Document()
    for word_file in word_files:
        doc = Document(word_file)
        for element in doc.element.body:
            merged_doc.element.body.append(element)
    merged_doc.save(output_file)

# Function to merge Excel files into a single Excel workbook
def merge_excel_files(excel_files, output_file):
    merged_workbook = px.Workbook()
    for excel_file in excel_files:
        workbook = px.load_workbook(excel_file)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            merged_workbook.create_sheet(title=sheet_name)
            merged_worksheet = merged_workbook[sheet_name]
            for row in worksheet.iter_rows():
                for cell in row:
                    merged_cell = merged_worksheet.cell(row=cell.row, column=cell.column, value=cell.value)
    merged_workbook.save(output_file)

if __name__ == "__main__":
    print("Welcome to DocMergePro - Your Document Merging Tool!")

    while True:
        print("\nMenu:")
        print("1. Merge PDF Files")
        print("2. Merge Word Files")
        print("3. Merge Excel Files")
        print("4. Quit")

        choice = input("Enter your choice (1/2/3/4): ").strip()

        if choice == '1':
            pdf_files = input("Enter the paths of the PDF files to merge (comma-separated): ").strip().split(',')
            output_file = input("Enter the path for the merged PDF output file: ").strip()
            merge_pdf_files(pdf_files, output_file)
            print(f"PDF files merged into '{output_file}'")
        elif choice == '2':
            word_files = input("Enter the paths of the Word files to merge (comma-separated): ").strip().split(',')
            output_file = input("Enter the path for the merged Word output file: ").strip()
            merge_word_files(word_files, output_file)
            print(f"Word files merged into '{output_file}'")
        elif choice == '3':
            excel_files = input("Enter the paths of the Excel files to merge (comma-separated): ").strip().split(',')
            output_file = input("Enter the path for the merged Excel output file: ").strip()
            merge_excel_files(excel_files, output_file)
            print(f"Excel files merged into '{output_file}'")
        elif choice == '4':
            break
        else:
            print("Invalid choice. Please select a valid option.")
